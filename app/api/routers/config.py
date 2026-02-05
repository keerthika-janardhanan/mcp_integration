from __future__ import annotations

from pathlib import Path
import os, subprocess, hashlib
from fastapi import APIRouter, HTTPException, Depends, UploadFile, File, Form
from pydantic import BaseModel, Field

from ..auth import jwt_required
from ..framework_resolver import resolve_framework_root
from ...services.config_service import update_test_manager_entry
from openpyxl import load_workbook


router = APIRouter(prefix="/config", tags=["config"], dependencies=[Depends(jwt_required)])


class UpdateTestManagerRequest(BaseModel):
    scenario: str
    datasheet: str
    referenceId: str
    idName: str
    frameworkRoot: str | None = Field(None, description="Root of the framework repo; resolves automatically if omitted")
    newDescription: str | None = Field(None, description="Optional new TestCaseDescription to apply")
    allowFreeformCreate: bool | None = Field(False, description="Permit creation of a new row even if scenario is not ID-like")
    execute: str | None = Field(None, description="Optional Execute value (defaults to 'Yes')")


@router.post("/update_test_manager")
async def update_test_manager(req: UpdateTestManagerRequest) -> dict:
    # Resolve framework root: accept local path or remote URL (clone if needed)
    repo_root: Path
    explicit = (req.frameworkRoot or "").strip() if req.frameworkRoot else ""
    if explicit:
        # Delegate full normalization + cloning to shared resolver for consistency
        try:
            repo_root = resolve_framework_root(explicit)
        except FileNotFoundError as exc:
            raise HTTPException(status_code=404, detail=str(exc)) from exc
        except Exception as exc:
            raise HTTPException(status_code=400, detail=f"Unable to resolve frameworkRoot: {exc}") from exc
    else:
        repo_root = resolve_framework_root()
    try:
        result = update_test_manager_entry(
            repo_root,
            scenario=req.scenario,
            execute_value=req.execute or "Yes",
            create_if_missing=True,
            datasheet=req.datasheet,
            reference_id=req.referenceId,
            id_name=req.idName,
            description_override=req.newDescription or None,
            allow_freeform_create=bool(req.allowFreeformCreate),
        )
    except RuntimeError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    except Exception as exc:  # pragma: no cover
        raise HTTPException(status_code=500, detail=str(exc)) from exc
    if not result:
        raise HTTPException(status_code=404, detail="testmanager.xlsx not found or invalid")
    return result


@router.get("/list_test_manager")
async def list_test_manager(frameworkRoot: str | None = None) -> dict:
    """Return rows from testmanager.xlsx with key columns for selection in UI."""
    repo_root: Path
    explicit = (frameworkRoot or "").strip() if frameworkRoot else ""
    if explicit:
        try:
            repo_root = resolve_framework_root(explicit)
        except FileNotFoundError as exc:
            raise HTTPException(status_code=404, detail=str(exc)) from exc
        except Exception as exc:
            raise HTTPException(status_code=400, detail=f"Unable to resolve frameworkRoot: {exc}") from exc
    else:
        repo_root = resolve_framework_root()

    from ...services.config_service import find_test_manager_path as _find
    tm = _find(repo_root)
    if not tm:
        return {"rows": []}
    try:
        wb = load_workbook(tm)
    except Exception:
        return {"rows": []}
    ws = wb.active
    # Map headers
    header = [str(c.value or "").strip() for c in ws[1]] if ws.max_row >= 1 else []
    def _idx(name: str) -> int | None:
        lname = name.lower()
        for i, h in enumerate(header):
            if lname == h.lower() or lname in h.lower():
                return i
        return None
    id_i = _idx("TestCaseID")
    desc_i = _idx("TestCaseDescription") or _idx("Description")
    exec_i = _idx("Execute")
    data_i = _idx("DatasheetName")
    ref_i = _idx("ReferenceID")
    name_i = _idx("IDName")

    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        rows.append({
            "TestCaseID": str((r[id_i] if id_i is not None and id_i < len(r) else "") or ""),
            "TestCaseDescription": str((r[desc_i] if desc_i is not None and desc_i < len(r) else "") or ""),
            "Execute": str((r[exec_i] if exec_i is not None and exec_i < len(r) else "") or ""),
            "DatasheetName": str((r[data_i] if data_i is not None and data_i < len(r) else "") or ""),
            "ReferenceID": str((r[ref_i] if ref_i is not None and ref_i < len(r) else "") or ""),
            "IDName": str((r[name_i] if name_i is not None and name_i < len(r) else "") or ""),
        })
    return {"rows": rows}


class RenameTestCaseIdRequest(BaseModel):
    oldTestCaseId: str
    newTestCaseId: str
    frameworkRoot: str | None = Field(None, description="Root of the framework repo; resolves automatically if omitted")


@router.post("/rename_test_case_id")
async def rename_test_case_id(req: RenameTestCaseIdRequest) -> dict:
    """Rename an existing TestCaseID in testmanager.xlsx"""
    # Resolve framework root
    repo_root: Path
    explicit = (req.frameworkRoot or "").strip() if req.frameworkRoot else ""
    if explicit:
        try:
            repo_root = resolve_framework_root(explicit)
        except FileNotFoundError as exc:
            raise HTTPException(status_code=404, detail=str(exc)) from exc
        except Exception as exc:
            raise HTTPException(status_code=400, detail=f"Unable to resolve frameworkRoot: {exc}") from exc
    else:
        repo_root = resolve_framework_root()
    
    from ...services.config_service import find_test_manager_path as _find
    tm_path = _find(repo_root)
    if not tm_path:
        raise HTTPException(status_code=404, detail="testmanager.xlsx not found")
    
    try:
        wb = load_workbook(tm_path)
        ws = wb.active
        
        # Find TestCaseID column
        header = [str(c.value or "").strip() for c in ws[1]] if ws.max_row >= 1 else []
        id_col_idx = None
        for i, h in enumerate(header):
            if h.lower() in ["testcaseid", "test case id", "testcase id"]:
                id_col_idx = i
                break
        
        if id_col_idx is None:
            raise HTTPException(status_code=400, detail="TestCaseID column not found in testmanager.xlsx")
        
        # Find and update the row
        found = False
        for row_idx in range(2, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=id_col_idx + 1)
            if str(cell.value or "").strip() == req.oldTestCaseId:
                cell.value = req.newTestCaseId
                found = True
                break
        
        if not found:
            raise HTTPException(status_code=404, detail=f"TestCaseID '{req.oldTestCaseId}' not found in testmanager.xlsx")
        
        # Save the workbook
        wb.save(tm_path)
        
        return {
            "success": True,
            "message": f"Successfully renamed '{req.oldTestCaseId}' to '{req.newTestCaseId}'",
            "oldTestCaseId": req.oldTestCaseId,
            "newTestCaseId": req.newTestCaseId
        }
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"Failed to rename TestCaseID: {exc}") from exc


@router.post("/upload_datasheet")
async def upload_datasheet(
    scenario: str = Form(...),
    frameworkRoot: str | None = Form(None),
    datasheetFile: UploadFile = File(...),
):
    """Upload a datasheet workbook or CSV into the framework's data/ directory.

    Returns JSON with saved relative path. The caller should then invoke /config/update_test_manager
    providing the datasheet file name and reference/idName fields.
    """
    # Resolve root same as update route (support remote URL)
    repo_root: Path
    explicit = (frameworkRoot or "").strip() if frameworkRoot else ""
    if explicit:
        try:
            repo_root = resolve_framework_root(explicit)
        except FileNotFoundError as exc:
            raise HTTPException(status_code=404, detail=str(exc)) from exc
        except Exception as exc:
            raise HTTPException(status_code=400, detail=f"Unable to resolve frameworkRoot: {exc}") from exc
    else:
        repo_root = resolve_framework_root()
    data_dir = repo_root / "data"
    try:
        data_dir.mkdir(parents=True, exist_ok=True)
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"Failed to create data directory: {exc}") from exc

    # Sanitize filename
    original_name = datasheetFile.filename or "datasheet.xlsx"
    safe_name = Path(original_name).name  # strip path parts
    if not safe_name:
        safe_name = "datasheet.xlsx"
    # Avoid overwriting by simple uniqueness suffix if file exists
    target_path = data_dir / safe_name
    counter = 1
    stem = target_path.stem
    suffix = target_path.suffix
    while target_path.exists():
        target_path = data_dir / f"{stem}_{counter}{suffix}"
        counter += 1
    try:
        contents = await datasheetFile.read()
        target_path.write_bytes(contents)
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(status_code=500, detail=f"Failed to save datasheet: {exc}") from exc

    rel = target_path.relative_to(repo_root).as_posix()
    return {"saved": rel, "filename": target_path.name, "scenario": scenario}


@router.get("/list_datasheets")
async def list_datasheets(frameworkRoot: str | None = None) -> dict:
    """List available datasheet files under the framework's data/ directory.

    Returns { files: ["data/FooData.xlsx", ...] } using POSIX-style relative paths.
    """
    # Resolve root (support remote URL resolution if provided)
    repo_root: Path
    explicit = (frameworkRoot or "").strip() if frameworkRoot else ""
    if explicit:
        try:
            repo_root = resolve_framework_root(explicit)
        except FileNotFoundError as exc:
            raise HTTPException(status_code=404, detail=str(exc)) from exc
        except Exception as exc:
            raise HTTPException(status_code=400, detail=f"Unable to resolve frameworkRoot: {exc}") from exc
    else:
        repo_root = resolve_framework_root()

    data_dir = repo_root / "data"
    if not data_dir.exists() or not data_dir.is_dir():
        return {"files": []}
    try:
        files = [
            p.name  # Return just the filename without data/ prefix
            for p in sorted(data_dir.glob("**/*"))
            if p.is_file() and p.suffix.lower() in {".xlsx", ".xls", ".csv"}
        ]
    except Exception as exc:  # pragma: no cover
        raise HTTPException(status_code=500, detail=f"Failed to enumerate datasheets: {exc}") from exc
    return {"files": files}
