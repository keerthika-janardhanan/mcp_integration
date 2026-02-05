from __future__ import annotations

from fastapi import APIRouter, HTTPException, Depends
import shutil
import logging
from pydantic import BaseModel, Field
from ..auth import jwt_required
from ..framework_resolver import resolve_framework_root
from pathlib import Path
from ..sse import _format_sse
from starlette.responses import StreamingResponse
from typing import AsyncGenerator
import re
from ...trial_spec_adapter import trial_env_overrides
from ...services.config_service import find_test_manager_path as _find_tm
try:
    from openpyxl import load_workbook  # type: ignore
except Exception:  # pragma: no cover
    load_workbook = None  # type: ignore


router = APIRouter(prefix="/agentic", tags=["agentic"])
logger = logging.getLogger(__name__)


class PreviewRequest(BaseModel):
    scenario: str


class PreviewResponse(BaseModel):
    preview: str


class RefineRequest(BaseModel):
    scenario: str
    previousPreview: str
    feedback: str


class PayloadRequest(BaseModel):
    scenario: str
    acceptedPreview: str


class FileItem(BaseModel):
    path: str
    content: str


class TestDataMapping(BaseModel):
    columnName: str
    occurrences: int
    actionType: str
    methods: list[str]


class PayloadResponse(BaseModel):
    locators: list[FileItem]
    pages: list[FileItem]
    tests: list[FileItem]
    testDataMapping: list[TestDataMapping]


def _unskip_tests_for_trial(source: str) -> tuple[str, int]:
    """Best-effort removal of declarative skips in test files for trial runs.

    We convert constructs like test.skip(name, fn) and test.describe.skip(name, fn)
    (and fixme variants) into active tests/blocks. This is applied ONLY for trial
    execution and never persisted to the repository.
    Returns (updated_source, replacements_count).
    """
    try:
        count = 0
        updated = source
        # 1) Replace describe-level skips/fixme at definition time
        for pat in (r"\btest\.describe\.skip\s*\(", r"\btest\.describe\.fixme\s*\("):
            updated, n = re.subn(pat, "test.describe(", updated)
            count += n
        # 2) Comment out runtime calls to test.skip()/test.fixme() inside bodies to avoid nested conversion
        def _comment_out_calls(src: str, name: str) -> tuple[str, int]:
            # Match start-of-line or after whitespace/semicolon until first closing paren and semicolon
            pattern = rf"(^|[;\s])test\.{name}\s*\([^;]*?\);"
            def repl(m: 're.Match[str]') -> str:
                prefix = m.group(1) or ""
                return prefix + f"// trial: removed test.{name}(...)"
            return re.subn(pattern, repl, src, flags=re.MULTILINE)
        updated, n1 = _comment_out_calls(updated, "skip")
        updated, n2 = _comment_out_calls(updated, "fixme")
        count += (n1 + n2)
        return updated, count
    except Exception:
        return source, 0


@router.post("/preview", response_model=PreviewResponse)
async def preview(req: PreviewRequest) -> PreviewResponse:
    try:
        from ...generators.agentic_script_agent import AgenticScriptAgent, FrameworkProfile
    except Exception as exc:  # pragma: no cover
        raise HTTPException(status_code=500, detail=f"Import failure: {exc}") from exc

    framework_root = resolve_framework_root()
    framework = FrameworkProfile.from_root(framework_root)
    agent = AgenticScriptAgent()
    try:
        context = agent.gather_context(req.scenario)
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"Context gathering failed: {exc}") from exc
    try:
        preview_text = agent.generate_preview(req.scenario, framework, context)
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"Preview generation failed: {exc}") from exc
    return PreviewResponse(preview=preview_text)


@router.post("/refine", response_model=PreviewResponse)
async def refine(req: RefineRequest) -> PreviewResponse:
    try:
        from ...generators.agentic_script_agent import AgenticScriptAgent, FrameworkProfile
    except Exception as exc:  # pragma: no cover
        raise HTTPException(status_code=500, detail=f"Import failure: {exc}") from exc
    
    framework_root = resolve_framework_root()
    framework = FrameworkProfile.from_root(framework_root)
    agent = AgenticScriptAgent()
    try:
        context = agent.gather_context(req.scenario)
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"Context gathering failed: {exc}") from exc
    try:
        refined_text = agent.refine_preview(req.scenario, framework, req.previousPreview, req.feedback, context)
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"Refine failed: {exc}") from exc
    return PreviewResponse(preview=refined_text)

@router.post("/payload", response_model=PayloadResponse)
async def payload(req: PayloadRequest) -> PayloadResponse:
    try:
        from ...generators.agentic_script_agent import AgenticScriptAgent, FrameworkProfile
    except Exception as exc:  # pragma: no cover
        raise HTTPException(status_code=500, detail=f"Import failure: {exc}") from exc
    
    framework_root = resolve_framework_root()
    framework = FrameworkProfile.from_root(framework_root)
    agent = AgenticScriptAgent()
    payload_dict = agent.generate_script_payload(req.scenario, framework, req.acceptedPreview)
    return PayloadResponse(
        locators=[FileItem(**f) for f in payload_dict.get("locators", [])],
        pages=[FileItem(**f) for f in payload_dict.get("pages", [])],
        tests=[FileItem(**f) for f in payload_dict.get("tests", [])],
        testDataMapping=[TestDataMapping(**m) for m in payload_dict.get("testDataMapping", [])],
    )


@router.post("/preview/stream")
async def preview_stream(req: PreviewRequest) -> StreamingResponse:
    """Stream progress events while generating an agentic preview.

    Events payload shape (JSON per SSE data frame):
      { "phase": "start" | "gather_context" | "context_ready" | "preview" | "done" | "error", ... }
    """
    try:
        from ...generators.agentic_script_agent import AgenticScriptAgent, FrameworkProfile
    except Exception as exc:  # pragma: no cover
        raise HTTPException(status_code=500, detail=f"Import failure: {exc}") from exc

    async def gen() -> AsyncGenerator[bytes, None]:
        try:
            yield _format_sse({"phase": "start"})
            framework_root = resolve_framework_root()
            framework = FrameworkProfile.from_root(framework_root)
            agent = AgenticScriptAgent()
            yield _format_sse({"phase": "gather_context"})
            context = agent.gather_context(req.scenario)
            flow_available = bool((context or {}).get("enriched_steps") or (context or {}).get("vector_steps"))
            yield _format_sse({"phase": "context_ready", "flow_available": flow_available})
            preview_text = agent.generate_preview(req.scenario, framework, context)
            yield _format_sse({"phase": "preview", "preview": preview_text})
            yield _format_sse({"phase": "done"})
        except Exception as exc:
            yield _format_sse({"phase": "error", "error": str(exc)})

    return StreamingResponse(gen(), media_type="text/event-stream")


@router.post("/payload/stream")
async def payload_stream(req: PayloadRequest) -> StreamingResponse:
    """Stream progress events while generating the agentic payload files.

    Event phases: start -> gather_context -> payload -> done (or error)
    """
    try:
        from ...generators.agentic_script_agent import AgenticScriptAgent, FrameworkProfile
    except Exception as exc:  # pragma: no cover
        raise HTTPException(status_code=500, detail=f"Import failure: {exc}") from exc

    async def gen() -> AsyncGenerator[bytes, None]:
        try:
            yield _format_sse({"phase": "start"})
            framework_root = resolve_framework_root()
            framework = FrameworkProfile.from_root(framework_root)
            agent = AgenticScriptAgent()
            yield _format_sse({"phase": "gather_context"})
            context = agent.gather_context(req.scenario)
            yield _format_sse({"phase": "context_ready", "flow_available": bool(context.get("vector_steps"))})
            payload_dict = agent.generate_script_payload(req.scenario, framework, req.acceptedPreview)
            # Only emit brief shapes to keep frames small
            summary = {
                "locators": len(payload_dict.get("locators", [])),
                "pages": len(payload_dict.get("pages", [])),
                "tests": len(payload_dict.get("tests", [])),
                "testDataMapping": len(payload_dict.get("testDataMapping", [])),
            }
            yield _format_sse({"phase": "payload", "summary": summary})
            yield _format_sse({"phase": "done"})
        except Exception as exc:
            yield _format_sse({"phase": "error", "error": str(exc)})

    return StreamingResponse(gen(), media_type="text/event-stream")


class PersistRequest(BaseModel):
    files: list[FileItem]
    frameworkRoot: str | None = None


@router.post("/persist", dependencies=[Depends(jwt_required)])
async def persist(req: PersistRequest) -> dict:
    try:
        from ...generators.agentic_script_agent import AgenticScriptAgent, FrameworkProfile
    except Exception as exc:  # pragma: no cover
        raise HTTPException(status_code=500, detail=f"Import failure: {exc}") from exc
    # Resolve provided frameworkRoot via shared resolver (supports remote URLs)
    framework_root = resolve_framework_root(req.frameworkRoot) if req.frameworkRoot else resolve_framework_root()
    framework = FrameworkProfile.from_root(framework_root)
    agent = AgenticScriptAgent()
    payload = {"locators": [], "pages": [], "tests": []}
    for f in req.files:
        # Group files by top-level folder name (locators/pages/tests)
        folder = (f.path.split("/")[0] or "").lower()
        if folder in payload:
            payload[folder].append({"path": f.path, "content": f.content})
        else:
            # default to tests if unknown
            payload["tests"].append({"path": f.path, "content": f.content})
    written = agent.persist_payload(framework, payload)
    rels = [str(p.relative_to(framework.root)).replace('\\', '/') for p in written]
    return {"written": rels}


class PushRequest(BaseModel):
    branch: str = Field("feature/agentic")
    message: str = Field("Add generated Playwright test")
    frameworkRoot: str | None = None


@router.post("/push", dependencies=[Depends(jwt_required)])
async def push(req: PushRequest) -> dict:
    try:
        from ...generators.agentic_script_agent import AgenticScriptAgent, FrameworkProfile
    except Exception as exc:  # pragma: no cover
        raise HTTPException(status_code=500, detail=f"Import failure: {exc}") from exc
    # Resolve provided frameworkRoot via shared resolver (supports remote URLs)
    framework_root = resolve_framework_root(req.frameworkRoot) if req.frameworkRoot else resolve_framework_root()
    framework = FrameworkProfile.from_root(framework_root)
    agent = AgenticScriptAgent()
    ok = agent.push_changes(framework, branch=req.branch, commit_msg=req.message)
    return {"success": bool(ok)}


class TrialRunRequest(BaseModel):
    testFileContent: str
    headed: bool = Field(True, description="Run browser in headed mode (defaults to true).")
    frameworkRoot: str | None = Field(None, description="Optional framework root to place temp spec inside tests dir")
    # Optional: before running, update testmanager.xlsx for this scenario
    scenario: str | None = Field(None, description="Scenario/TestCase identifier to enable in testmanager.xlsx")
    updateTestManager: bool = Field(False, description="If true and scenario provided, set Execute='Yes' and update datasheet mapping")
    datasheet: str | None = Field(None, description="Datasheet file name to write into testmanager.xlsx (optional)")
    referenceId: str | None = Field(None, description="ReferenceID value to write into testmanager.xlsx (optional)")
    referenceIds: list[str] | None = Field(None, description="Optional list of ReferenceIDs to run sequentially (max 3) for generated streaming runs. If provided, takes precedence over referenceId.")
    idName: str | None = Field(None, description="IDName (column name) to write into testmanager.xlsx (optional)")


class TrialRunResponse(BaseModel):
    success: bool
    logs: str
    updateInfo: dict | None = None


@router.post("/trial-run", response_model=TrialRunResponse)
async def trial_run(req: TrialRunRequest) -> TrialRunResponse:
    """Execute a temporary Playwright test file. If frameworkRoot provided, place spec inside its tests dir to honor config."""
    try:
        from ...executor import run_trial, run_trial_in_framework
        from ..framework_resolver import resolve_framework_root
        from pathlib import Path as _P
    except Exception as exc:  # pragma: no cover
        raise HTTPException(status_code=500, detail=f"Import failure: {exc}") from exc

    upd_info = None
    if req.frameworkRoot:
        try:
            # Allow remote git URLs by delegating to resolver first
            try:
                root = resolve_framework_root(req.frameworkRoot)
            except Exception:
                # Fallback to local path resolution if it's not a URL or resolver failed
                root = _P(req.frameworkRoot).expanduser().resolve()
            if not root.exists():
                raise HTTPException(status_code=404, detail=f"frameworkRoot not found: {root}")
            # Optionally update testmanager.xlsx prior to running the trial
            if req.updateTestManager and (req.scenario or "").strip():
                try:
                    from ...services.config_service import update_test_manager_entry as _upd
                    upd = _upd(
                        root,
                        scenario=(req.scenario or "").strip(),
                        execute_value="Yes",
                        create_if_missing=True,
                        datasheet=(req.datasheet or None),
                        reference_id=(req.referenceId or None),
                        id_name=(req.idName or None),
                    )
                    upd_info = upd or None
                    if not upd:
                        # testmanager.xlsx missing or invalid
                        raise HTTPException(status_code=404, detail="testmanager.xlsx not found or invalid")
                except HTTPException:
                    raise
                except Exception as exc:
                    raise HTTPException(status_code=500, detail=f"Failed updating testmanager.xlsx: {exc}") from exc

            # Auto-create minimal stubs for missing page object imports to prevent module not found errors
            try:
                missing_created = []
                # Find import lines like: import X from "../pages/SomePage.ts";
                pattern = re.compile(r'import\s+[^;]*?from\s+"(\.\./pages/[^"\n]+\.ts)"')
                for match in pattern.finditer(req.testFileContent):
                    rel_path = match.group(1)
                    target_path = (root / rel_path).resolve()
                    # Security: ensure within root
                    try:
                        target_path.relative_to(root)
                    except ValueError:
                        continue
                    if not target_path.exists():
                        target_path.parent.mkdir(parents=True, exist_ok=True)
                        # Derive class/interface name crudely from filename
                        base_name = target_path.stem
                        class_name = re.sub(r'[^A-Za-z0-9]', '', base_name.title()) or 'PageObject'
                        stub = (
                            f"// Auto-generated stub to unblock trial run for {base_name}\n"
                            f"export default class {class_name} {{\n"
                            f"  constructor(page) {{ this.page = page; }}\n"
                            f"  async placeholder() {{ /* implement actions */ }}\n"
                            f"}}\n"
                        )
                        try:
                            target_path.write_text(stub, encoding='utf-8')
                            missing_created.append(str(target_path.relative_to(root)))
                        except Exception:
                            pass
                # Optionally could log created stubs; for now silent
            except Exception:
                pass
            # Temporarily unskip tests for trial runs only (not persisted)
            content, replaced = _unskip_tests_for_trial(req.testFileContent)
            env_overrides = trial_env_overrides(root, case_id=(req.scenario or None))
            # Compose a non-sensitive banner indicating chosen trial credentials
            def _mask_pw(pw: str | None) -> str:
                if not pw:
                    return ""
                if len(pw) <= 2:
                    return "***"
                return ("*" * (len(pw) - 2)) + pw[-2:]
            user = env_overrides.get("USERID") or env_overrides.get("USERNAME") or env_overrides.get("TRIAL_USERNAME") or env_overrides.get("EMAIL") or ""
            pw = env_overrides.get("PASSWORD") or env_overrides.get("TRIAL_PASSWORD") or ""
            base = env_overrides.get("BASE_URL") or env_overrides.get("URL") or env_overrides.get("TRIAL_BASE_URL") or env_overrides.get("TRIAL_URL") or ""
            banner = "[trial-creds] username=" + (user or "<empty>") + ", password=" + _mask_pw(pw) + (", base_url=" + base if base else "") + "\n"
            # Best-effort clean previous results to avoid artifact collisions
            try:
                results_dir = root / 'test-results'
                if results_dir.exists():
                    shutil.rmtree(results_dir, ignore_errors=True)
            except Exception:
                pass
            success, logs = run_trial_in_framework(content, root, headed=req.headed, env_overrides=env_overrides)
            logger.info(banner.strip())
            logs = banner + logs
            if replaced:
                logs = f"[trial-note] Unskipped {replaced} skipped/fixme declarations for this run.\n" + logs
        except HTTPException:
            raise
        except Exception as exc:
            raise HTTPException(status_code=500, detail=f"Framework trial failure: {exc}") from exc
    else:
        content, replaced = _unskip_tests_for_trial(req.testFileContent)
        # Use resolved framework root (default) for credentials fallback; safe if not present
        default_root = resolve_framework_root()
        env_overrides = trial_env_overrides(default_root, case_id=(req.scenario or None))
        def _mask_pw(pw: str | None) -> str:
            if not pw:
                return ""
            if len(pw) <= 2:
                return "***"
            return ("*" * (len(pw) - 2)) + pw[-2:]
        user = env_overrides.get("USERID") or env_overrides.get("USERNAME") or env_overrides.get("TRIAL_USERNAME") or env_overrides.get("EMAIL") or ""
        pw = env_overrides.get("PASSWORD") or env_overrides.get("TRIAL_PASSWORD") or ""
        base = env_overrides.get("BASE_URL") or env_overrides.get("URL") or env_overrides.get("TRIAL_BASE_URL") or env_overrides.get("TRIAL_URL") or ""
        banner = "[trial-creds] username=" + (user or "<empty>") + ", password=" + _mask_pw(pw) + (", base_url=" + base if base else "") + "\n"
        success, logs = run_trial(content, headed=req.headed, env_overrides=env_overrides)
        logger.info(banner.strip())
        logs = banner + logs
        if replaced:
            logs = f"[trial-note] Unskipped {replaced} skipped/fixme declarations for this run.\n" + logs
    return TrialRunResponse(success=bool(success), logs=logs, updateInfo=upd_info)


class TrialRunExistingRequest(BaseModel):
    testFilePath: str = Field(..., description="Relative path to existing test file inside framework repo")
    headed: bool = Field(True)
    frameworkRoot: str | None = Field(None, description="Optional explicit framework root; auto-resolved if omitted")
    # Optional: update testmanager before running
    scenario: str | None = Field(None, description="Scenario/TestCase identifier to enable in testmanager.xlsx")
    updateTestManager: bool = Field(False, description="If true and scenario provided, set Execute='Yes' and update datasheet mapping")
    datasheet: str | None = Field(None, description="Datasheet file name to write into testmanager.xlsx (optional)")
    referenceId: str | None = Field(None, description="ReferenceID value to write into testmanager.xlsx (optional). Supports comma-separated for parallel runs.")
    referenceIds: list[str] | None = Field(None, description="Optional list of ReferenceIDs to run in parallel (max 3). If provided, takes precedence over referenceId.")
    idName: str | None = Field(None, description="IDName (column name) to write into testmanager.xlsx (optional)")


@router.post("/trial-run-existing", response_model=TrialRunResponse)
async def trial_run_existing(req: TrialRunExistingRequest) -> TrialRunResponse:
    """Execute an existing test file from the framework repository.

    Reads the file content and delegates to run_trial (temp spec execution) so we don't mutate repo.
    """
    logger.info(f"[TrialRunExisting] Request received:")
    logger.info(f"[TrialRunExisting]   testFilePath={req.testFilePath}")
    logger.info(f"[TrialRunExisting]   scenario={req.scenario}")
    logger.info(f"[TrialRunExisting]   updateTestManager={req.updateTestManager}")
    logger.info(f"[TrialRunExisting]   headed={req.headed}")
    
    try:
        from ...executor import run_trial_in_framework
        from ..framework_resolver import resolve_framework_root
    except Exception as exc:  # pragma: no cover
        raise HTTPException(status_code=500, detail=f"Import failure: {exc}") from exc
    # Resolve provided frameworkRoot via shared resolver (supports remote URLs)
    root = resolve_framework_root(req.frameworkRoot) if req.frameworkRoot else resolve_framework_root()
    
    # Optionally update testmanager.xlsx prior to running the trial
    upd_info = None
    if req.updateTestManager and (req.scenario or "").strip():
        logger.info(f"[TrialRunExisting] Updating testmanager.xlsx for scenario: '{req.scenario}'")
        try:
            from ...services.config_service import update_test_manager_entry as _upd
            upd_info = _upd(
                root,
                scenario=(req.scenario or "").strip(),
                execute_value="Yes",
                create_if_missing=True,
                datasheet=(req.datasheet or None),
                # If multiple reference IDs supplied, don't persist a specific one into Excel to avoid races
                reference_id=(None if (req.referenceIds and len(req.referenceIds) > 1) else (req.referenceId or None)),
                id_name=(req.idName or None),
            )
            logger.info(f"[TrialRunExisting] TestManager update result: {upd_info}")
        except Exception as exc:
            # Non-fatal: proceed with trial even if update fails
            logger.error(f"[TrialRunExisting] TestManager update failed: {exc}", exc_info=True)
            pass
    else:
        logger.info(f"[TrialRunExisting] Skipping testmanager update (updateTestManager={req.updateTestManager}, scenario='{req.scenario}')")
    target = (root / req.testFilePath).resolve()
    try:
        if target.is_dir():
            raise HTTPException(status_code=400, detail="testFilePath points to a directory")
        if not target.exists():
            raise HTTPException(status_code=404, detail=f"Test file not found: {req.testFilePath}")
        # Prevent path escape
        root_resolved = root.resolve()
        if root_resolved not in target.parents and target != root_resolved:
            raise HTTPException(status_code=400, detail="testFilePath escapes framework root")
        content = target.read_text(encoding="utf-8", errors="replace")
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"Failed reading file: {exc}") from exc
    
    # Adapt content for trial (disable tracing, isolate outputs, env-driven overrides)
    try:
        from ...trial_spec_adapter import adapt_spec_content_for_trial
        adapted_content, was_adapted = adapt_spec_content_for_trial(content, root)
        if was_adapted:
            content = adapted_content
            logger.info("[TrialRunExisting] Applied trial spec adaptations (trace off, outputDir isolation, env overrides)")
    except Exception as _adapt_exc:
        logger.warning(f"[TrialRunExisting] Spec adaptation skipped due to error: {_adapt_exc}")

    env_overrides = trial_env_overrides(root, case_id=(req.scenario or None), spec_path=target)
    
    def _mask_pw(pw: str | None) -> str:
        if not pw:
            return ""
        if len(pw) <= 2:
            return "***"
        return ("*" * (len(pw) - 2)) + pw[-2:]
    
    user = env_overrides.get("USERID") or env_overrides.get("USERNAME") or env_overrides.get("TRIAL_USERNAME") or env_overrides.get("EMAIL") or ""
    pw = env_overrides.get("PASSWORD") or env_overrides.get("TRIAL_PASSWORD") or ""
    base = env_overrides.get("BASE_URL") or env_overrides.get("URL") or env_overrides.get("TRIAL_BASE_URL") or env_overrides.get("TRIAL_URL") or ""
    banner = "[trial-creds] username=" + (user or "<empty>") + ", password=" + _mask_pw(pw) + (", base_url=" + base if base else "") + "\n"
    
    # For trial runs, inject credentials via environment variables (not .env file)
    # The test code will access them via process.env.USERID, process.env.PASSWORD, etc.
    logger.info(f"[TrialRunExisting] Injecting trial credentials into environment")
    logger.info(f"[TrialRunExisting] Credentials: USERID={user[:20] if user else '<empty>'}, PASSWORD={'***' if pw else '<empty>'}, BASE_URL={base[:30] if base else '<empty>'}")
    
    # Add Node.js to PATH for subprocess
    import os as _os
    nodejs_path = r"C:\Program Files\nodejs"
    if nodejs_path not in env_overrides.get("PATH", ""):
        current_path = env_overrides.get("PATH", _os.environ.get("PATH", ""))
        env_overrides["PATH"] = nodejs_path + _os.pathsep + current_path
        logger.info(f"[TrialRunExisting] Added Node.js to PATH: {nodejs_path}")
    
    # Best-effort clean previous results to avoid artifact collisions across sequential/parallel runs
    try:
        results_dir = root / 'test-results'
        if results_dir.exists():
            shutil.rmtree(results_dir, ignore_errors=True)
            logger.info(f"[TrialRunExisting] Cleaned previous results at {results_dir}")
    except Exception as _cleanup_exc:
        logger.warning(f"[TrialRunExisting] Failed to clean test-results: {_cleanup_exc}")
    
    # Prepare trial execution using temp-spec runner so we can unskip tests
    import os
    from concurrent.futures import ThreadPoolExecutor, as_completed
    try:
        from ...executor import run_trial_in_framework  # temp spec inside framework tests dir
    except Exception as exc:  # pragma: no cover
        raise HTTPException(status_code=500, detail=f"Import failure: {exc}") from exc
    # Unskip tests for trial-only execution
    content, replaced = _unskip_tests_for_trial(content)

    # Helper: read ReferenceID/IDName from testmanager.xlsx for the provided scenario
    def _read_refs_from_excel(root_path, scenario: str) -> tuple[str | None, str | None]:
        try:
            if not scenario:
                return None, None
            tm = _find_tm(root_path)
            if not tm or not load_workbook:
                return None, None
            wb = load_workbook(tm)
            ws = wb.active
            header = [str(c.value or "").strip() for c in ws[1]] if ws.max_row >= 1 else []
            def _idx(name: str) -> int | None:
                lname = name.lower()
                for i, h in enumerate(header):
                    if lname == h.lower() or lname in h.lower():
                        return i
                return None
            id_i = _idx("TestCaseID")
            ref_i = _idx("ReferenceID")
            name_i = _idx("IDName")
            if id_i is None:
                return None, None
            for r in ws.iter_rows(min_row=2, values_only=True):
                case_id = str((r[id_i] if id_i is not None and id_i < len(r) else "") or "").strip()
                if not case_id:
                    continue
                if case_id.lower() == scenario.strip().lower():
                    ref_val = str((r[ref_i] if ref_i is not None and ref_i < len(r) else "") or "").strip() if ref_i is not None else ""
                    id_name_val = str((r[name_i] if name_i is not None and name_i < len(r) else "") or "").strip() if name_i is not None else ""
                    return (ref_val or None), (id_name_val or None)
        except Exception:
            return None, None
        return None, None

    # Determine ReferenceIDs to run
    ref_ids: list[str] = []
    if req.referenceIds and isinstance(req.referenceIds, list) and len(req.referenceIds) > 0:
        ref_ids = [str(r).strip() for r in req.referenceIds if str(r).strip()]
    elif req.referenceId and "," in req.referenceId:
        ref_ids = [r.strip() for r in (req.referenceId or "").split(",") if r.strip()]
    elif req.referenceId:
        ref_ids = [str(req.referenceId).strip()]

    # If not provided via API, fall back to Excel row for the scenario
    excel_ref_val: str | None = None
    excel_idname: str | None = None
    effective_id_name = req.idName
    if not ref_ids and (req.scenario or "").strip():
        excel_ref_val, excel_idname = _read_refs_from_excel(root, (req.scenario or "").strip())
        if excel_ref_val:
            # Support comma, semicolon, whitespace, and newline-separated values
            import re as _re
            parts = [p.strip() for p in _re.split(r"[,;\s]+", str(excel_ref_val)) if p and str(p).strip()]
            if parts:
                ref_ids = parts
        if not effective_id_name and excel_idname:
            effective_id_name = excel_idname
    if not effective_id_name:
        effective_id_name = req.idName

    # If multiple reference IDs provided, run in parallel (max 3). Otherwise, single run.
    if len(ref_ids) <= 1:
        env = os.environ.copy()
        env.update(env_overrides)
        if ref_ids:
            env.update({
                "REFERENCE_ID": ref_ids[0],
                "DATA_REFERENCE_ID": ref_ids[0],
            })
            if effective_id_name:
                env.update({
                    "ID_NAME": effective_id_name,
                    "DATA_ID_NAME": effective_id_name,
                })
        success, logs = run_trial_in_framework(content, root, headed=req.headed, env_overrides=env)
        logs = (f"[trial-note] Unskipped tests for this run.\n" if replaced else "") + banner + logs
        return TrialRunResponse(success=bool(success), logs=logs, updateInfo=upd_info)

    # Parallel path
    max_workers = min(3, len(ref_ids))
    logger.info(f"[TrialRunExisting] Running parallel executions for {len(ref_ids)} ReferenceIDs (max_workers={max_workers})")

    def _run_for_ref(ref: str) -> tuple[str, bool, str]:
        env = os.environ.copy()
        env.update(env_overrides)
        env.update({
            "REFERENCE_ID": ref,
            "DATA_REFERENCE_ID": ref,
        })
        if effective_id_name:
            env.update({
                "ID_NAME": effective_id_name,
                "DATA_ID_NAME": effective_id_name,
            })
        try:
            ok, logs = run_trial_in_framework(content, root, headed=req.headed, env_overrides=env)
            combined = f"[reference:{ref}]\n" + ((f"[trial-note] Unskipped tests for this run.\n" if replaced else "") + banner + logs)
            return ref, bool(ok), combined
        except Exception as exec_exc:
            logger.error(f"[TrialRunExisting] Parallel execution failed for {ref}: {exec_exc}", exc_info=True)
            return ref, False, f"[reference:{ref}]\n" + banner + f"Execution error: {exec_exc}"

    results: list[tuple[str, bool, str]] = []
    with ThreadPoolExecutor(max_workers=max_workers) as pool:
        futures = {pool.submit(_run_for_ref, ref): ref for ref in ref_ids}
        for fut in as_completed(futures):
            try:
                results.append(fut.result())
            except Exception as e:  # pragma: no cover
                r = futures[fut]
                logger.error(f"[TrialRunExisting] Worker crashed for {r}: {e}")
                results.append((r, False, f"[reference:{r}]\n" + banner + f"Execution error: {e}"))

    # Summarize
    results.sort(key=lambda t: ref_ids.index(t[0]))
    all_success = all(s for (_, s, _) in results) if results else False
    summary_lines = [f"{('PASS' if s else 'FAIL')} - {p}" for (p, s, _) in results]
    combined_logs = "\n\n".join(log for (_, _, log) in results)
    summary = "Parallel run summary (ReferenceIDs)\n" + "\n".join(summary_lines) + "\n\n" + combined_logs
    return TrialRunResponse(success=all_success, logs=summary, updateInfo=upd_info)


class KeywordInspectRequest(BaseModel):
    keyword: str = Field(..., description="Scenario keyword to inspect against repo and refined recorder flows")
    repoPath: str = Field(..., description="Framework repository path or git URL")
    branch: str | None = Field(None, description="Branch to use (optional if embedded in URL)")
    maxAssets: int = Field(5, ge=1, le=25, description="Maximum existing framework assets to return")


class ExistingAsset(BaseModel):
    path: str
    snippet: str
    isTest: bool = False
    relevance: int | None = None


class RefinedRecorderFlow(BaseModel):
    sourceSession: str | None = None
    steps: list[dict] = []
    stabilityWarnings: list[str] = []


class VectorContext(BaseModel):
    flowAvailable: bool
    vectorStepsCount: int


class KeywordInspectResponse(BaseModel):
    keyword: str
    existingAssets: list[ExistingAsset]
    refinedRecorderFlow: RefinedRecorderFlow | None
    vectorContext: VectorContext
    status: str
    messages: list[str]


@router.post("/keyword-inspect", response_model=KeywordInspectResponse)
async def keyword_inspect(req: KeywordInspectRequest) -> KeywordInspectResponse:
    """Inspect a keyword against the framework repo and refined recorder/vector flows.

    Flow:
      1. Create/clone framework repository locally under framework_repos/
      2. Search for keyword in tests/*.spec.ts files (existing scripts)
      3. Search for refined recorder flows in vector DB
      4. Return both existing scripts and refined flows
    """
    try:
        from ...generators.agentic_script_agent import AgenticScriptAgent, FrameworkProfile
        from ..framework_resolver import resolve_framework_root
    except Exception as exc:  # pragma: no cover
        raise HTTPException(status_code=500, detail=f"Import failure: {exc}") from exc

    # Validate inputs
    keyword = (req.keyword or "").strip()
    if not keyword:
        raise HTTPException(status_code=400, detail="Keyword is required")
    repo_input = (req.repoPath or "").strip()
    if not repo_input:
        raise HTTPException(status_code=400, detail="repoPath is mandatory")

    messages = []
    
    try:
        # Step 1: Resolve/create framework repository locally
        # This will clone remote repos to framework_repos/<hash> or create default if nothing provided
        framework_root: Path
        try:
            framework_root = resolve_framework_root(repo_input)
            messages.append(f"Framework repository resolved: {framework_root}")
        except FileNotFoundError as exc:
            raise HTTPException(status_code=404, detail=str(exc)) from exc
        except Exception as exc:
            raise HTTPException(status_code=400, detail=f"Unable to resolve repository: {exc}") from exc

        framework = FrameworkProfile.from_root(framework_root)
        agent = AgenticScriptAgent()

        # Step 2: Search for existing test files with keyword
        existing_assets = []
        
        logger.info(f"[KeywordInspect] Framework root: {framework_root}")
        logger.info(f"[KeywordInspect] Searching for keyword: '{keyword}'")
        
        # Search in multiple common test directories
        search_dirs = [
            framework_root / "tests",
            framework_root / "test", 
            framework_root / "e2e",
            framework_root / "specs",
            framework_root / "__tests__",
            framework_root  # Fallback: search entire repo
        ]
        
        # Also check if there's a specific tests directory from framework profile
        try:
            framework = FrameworkProfile.from_root(framework_root)
            if framework.tests_dir and (framework_root / framework.tests_dir).exists():
                tests_from_profile = framework_root / framework.tests_dir
                if tests_from_profile not in search_dirs:
                    search_dirs.insert(0, tests_from_profile)
                    logger.info(f"[KeywordInspect] Added tests dir from framework profile: {tests_from_profile}")
        except Exception as e:
            logger.warning(f"[KeywordInspect] Could not load framework profile: {e}")
        
        spec_files = []
        searched_paths = []
        
        for search_dir in search_dirs:
            if search_dir.exists():
                logger.info(f"[KeywordInspect] Searching in: {search_dir}")
                searched_paths.append(str(search_dir.relative_to(framework_root)) if search_dir != framework_root else ".")
                
                # Search for spec files
                found_in_dir = list(search_dir.glob("**/*.spec.ts")) + list(search_dir.glob("**/*.test.ts"))
                
                # Avoid duplicates
                for f in found_in_dir:
                    if f not in spec_files:
                        spec_files.append(f)
                
                logger.info(f"[KeywordInspect] Found {len(found_in_dir)} files in {search_dir.name}")
        
        messages.append(f"Searched in: {', '.join(searched_paths)}")
        messages.append(f"Found {len(spec_files)} test files to search")
        logger.info(f"[KeywordInspect] Total unique spec files found: {len(spec_files)}")
        
        if len(spec_files) > 0:
            # Log all file paths found
            for sf in spec_files:
                logger.info(f"[KeywordInspect] File in list: {sf.relative_to(framework_root)}")
            
            for spec_file in spec_files:
                logger.info(f"[KeywordInspect] ========================================")
                logger.info(f"[KeywordInspect] Checking file: {spec_file.name}")
                logger.info(f"[KeywordInspect] Full path: {spec_file}")
                try:
                    content = spec_file.read_text(encoding='utf-8')
                    logger.info(f"[KeywordInspect] File size: {len(content)} chars, {len(content.split(chr(10)))} lines")
                    
                    # More flexible keyword matching:
                    # 1. Match against filename
                    # 2. Direct keyword match in content (case-insensitive)
                    # 3. Keyword with spaces/underscores/hyphens normalized
                    keyword_lower = keyword.lower()
                    keyword_normalized = keyword_lower.replace(' ', '').replace('_', '').replace('-', '')
                    
                    # Check filename first
                    filename_lower = spec_file.name.lower()
                    filename_normalized = filename_lower.replace(' ', '').replace('_', '').replace('-', '').replace('.spec.ts', '').replace('.test.ts', '')
                    
                    content_lower = content.lower()
                    content_normalized = content_lower.replace(' ', '').replace('_', '').replace('-', '')
                    
                    logger.info(f"[KeywordInspect] Keyword (original): '{keyword}'")
                    logger.info(f"[KeywordInspect] Keyword (lower): '{keyword_lower}'")
                    logger.info(f"[KeywordInspect] Keyword (normalized): '{keyword_normalized}'")
                    logger.info(f"[KeywordInspect] Filename (normalized): '{filename_normalized}'")
                    logger.info(f"[KeywordInspect] Content preview (first 200 chars): {content[:200]}")
                    
                    match_found = False
                    match_type = None
                    
                    # Check filename match
                    if keyword_lower in filename_lower or keyword_normalized in filename_normalized:
                        match_found = True
                        match_type = "filename"
                        logger.info(f"[KeywordInspect] MATCH FOUND: Filename match")
                    # Check content match
                    elif keyword_lower in content_lower:
                        match_found = True
                        match_type = "direct"
                        logger.info(f"[KeywordInspect] MATCH FOUND: Direct match")
                    elif keyword_normalized in content_normalized:
                        match_found = True
                        match_type = "normalized"
                        logger.info(f"[KeywordInspect] MATCH FOUND: Normalized match")
                    else:
                        logger.info(f"[KeywordInspect] NO MATCH: Keyword not found in filename or content")
                    
                    logger.info(f"[KeywordInspect] Match found: {match_found}, type: {match_type}")
                    
                    if match_found:
                        # Extract a snippet around the keyword
                        lines = content.split('\n')
                        matching_lines = []
                        
                        # For filename matches, show the beginning of the file
                        if match_type == "filename":
                            # Show first 5 lines as snippet
                            snippet = '\n'.join(lines[:5])
                            matching_lines = [0]  # Mark as having at least 1 match
                        else:
                            # For content matches, find matching lines
                            for i, line in enumerate(lines):
                                line_lower = line.lower()
                                line_normalized = line_lower.replace(' ', '').replace('_', '').replace('-', '')
                                if keyword_lower in line_lower or keyword_normalized in line_normalized:
                                    matching_lines.append(i)
                            
                            if matching_lines:
                                # Get context around first match
                                match_idx = matching_lines[0]
                                start = max(0, match_idx - 2)
                                end = min(len(lines), match_idx + 3)
                                snippet = '\n'.join(lines[start:end])
                            else:
                                # Fallback: show first 5 lines
                                snippet = '\n'.join(lines[:5])
                                matching_lines = [0]
                        
                        logger.info(f"[KeywordInspect] Found keyword in {len(matching_lines)} lines (match type: {match_type})")
                        logger.info(f"[KeywordInspect] Found keyword in {len(matching_lines)} lines (match type: {match_type})")
                        
                        if matching_lines or match_type == "filename":
                            relative_path = str(spec_file.relative_to(framework_root)).replace('\\', '/')
                            existing_assets.append(ExistingAsset(
                                path=relative_path,
                                snippet=snippet[:500],  # Increased snippet length
                                isTest=True,
                                relevance=len(matching_lines) if matching_lines else 1
                            ))
                            msg = f"Found keyword in {relative_path} ({len(matching_lines)} occurrences, match type: {match_type})"
                            messages.append(msg)
                            logger.info(f"[KeywordInspect] {msg}")
                except Exception as e:
                    error_msg = f"Error reading {spec_file.name}: {str(e)}"
                    messages.append(error_msg)
                    logger.error(f"[KeywordInspect] {error_msg}")
        else:
            msg = f"No .spec.ts or .test.ts files found in repository"
            messages.append(msg)
            logger.warning(f"[KeywordInspect] {msg}")

        # Step 3: Search for refined recorder flows in vector DB
        vector_context = VectorContext(flowAvailable=False, vectorStepsCount=0)
        refined_flow = None
        
        try:
            context = agent.gather_context(keyword)
            vector_steps = context.get("vector_steps", [])
            
            if vector_steps:
                vector_context.flowAvailable = True
                vector_context.vectorStepsCount = len(vector_steps)

                # Default: return all refined steps. Allow optional cap via REFINED_FLOW_PREVIEW_MAX_STEPS.
                try:
                    import os as _os
                    _limit_raw = _os.getenv("REFINED_FLOW_PREVIEW_MAX_STEPS")
                    if _limit_raw is None or str(_limit_raw).strip() == "":
                        _limit = None
                    else:
                        _raw = str(_limit_raw).strip().lower()
                        if _raw in {"all", "unlimited", "none"}:
                            _limit = None
                        else:
                            _n = int(_raw)
                            _limit = None if _n <= 0 else _n
                except Exception:
                    _limit = None

                _steps = vector_steps if _limit is None else vector_steps[: max(1, _limit)]
                refined_flow = RefinedRecorderFlow(
                    sourceSession=context.get("session_id"),
                    steps=_steps,
                    stabilityWarnings=[]
                )
                messages.append(f"Found refined recorder flow with {len(vector_steps)} steps")
            else:
                messages.append("No refined recorder flow found in vector DB")
        except Exception as e:
            messages.append(f"Error gathering vector context: {str(e)}")

        # Sort existing assets by relevance (most matches first)
        existing_assets.sort(key=lambda x: x.relevance or 0, reverse=True)
        existing_assets = existing_assets[:req.maxAssets]

        return KeywordInspectResponse(
            keyword=keyword,
            existingAssets=existing_assets,
            refinedRecorderFlow=refined_flow,
            vectorContext=vector_context,
            status="success",
            messages=messages
        )
    except HTTPException:
        # Preserve intended HTTP status for validation/git errors
        raise
    except Exception as fatal_exc:
        # Final catch-all: return structured error response instead of 500
        return KeywordInspectResponse(
            keyword=(req.keyword or "").strip(),
            existingAssets=[],
            refinedRecorderFlow=None,
            vectorContext=VectorContext(flowAvailable=False, vectorStepsCount=0),
            status="error",
            messages=[f"keyword-inspect failed: {type(fatal_exc).__name__}: {fatal_exc}"],
        )

@router.post("/trial-run/stream")
async def trial_run_stream(req: TrialRunRequest) -> StreamingResponse:
    """Stream real-time execution logs of a temporary Playwright test via SSE.

    Phases: start -> running -> chunk (repeated) -> done OR error
    Each chunk frame contains {"phase": "chunk", "data": "..."}
    Final frame includes {"phase": "done", "success": bool}
    """
    try:
        import asyncio
        import tempfile, os, subprocess
        from pathlib import Path as _P
        from ...executor import _resolve_playwright_command, _detect_test_dir
        from ..framework_resolver import resolve_framework_root as _resolve_root
    except Exception as exc:  # pragma: no cover
        raise HTTPException(status_code=500, detail=f"Import failure: {exc}") from exc

    async def gen() -> AsyncGenerator[bytes, None]:
        try:
            logger.info(f"[TrialRunStream] Starting trial run stream - headed={req.headed}, frameworkRoot={req.frameworkRoot}")
            yield _format_sse({"phase": "start"})
            tmp_path = None
            cwd = None
            cmd = None
            # If a frameworkRoot is specified, write inside its detected testDir so Playwright config applies.
            # Helper: Excel fallback for ReferenceIDs if not provided
            def _excel_refs(root_path, scenario: str) -> tuple[list[str], str | None]:
                refs: list[str] = []
                id_name: str | None = None
                if not scenario or not load_workbook:
                    return refs, id_name
                try:
                    tm = _find_tm(root_path)
                    if not tm:
                        return refs, id_name
                    wb = load_workbook(tm)
                    ws = wb.active
                    header = [str(c.value or "").strip() for c in ws[1]] if ws.max_row >= 1 else []
                    def _idx(name: str) -> int | None:
                        lname = name.lower()
                        for i, h in enumerate(header):
                            if lname == h.lower() or lname in h.lower():
                                return i
                        return None
                    id_i = _idx("TestCaseID")
                    ref_i = _idx("ReferenceID")
                    name_i = _idx("IDName")
                    if id_i is None:
                        return refs, id_name
                    for r in ws.iter_rows(min_row=2, values_only=True):
                        case_id = str((r[id_i] if id_i is not None and id_i < len(r) else "") or "").strip()
                        if not case_id:
                            continue
                        if case_id.lower() == scenario.strip().lower():
                            raw_ref = str((r[ref_i] if ref_i is not None and ref_i < len(r) else "") or "").strip()
                            raw_name = str((r[name_i] if name_i is not None and name_i < len(r) else "") or "").strip()
                            if raw_ref:
                                import re as _re
                                parts = [p.strip() for p in _re.split(r"[,;\s]+", raw_ref) if p.strip()]
                                refs = parts
                            if raw_name:
                                id_name = raw_name
                            break
                except Exception:
                    return refs, id_name
                return refs, id_name

            if req.frameworkRoot:
                logger.info(f"[TrialRunStream] Using frameworkRoot: {req.frameworkRoot}")
                try:
                    # Resolve local path or remote git URL consistently
                    root = _resolve_root(req.frameworkRoot)
                    logger.info(f"[TrialRunStream] Resolved root: {root}")
                except Exception as exc:
                    logger.error(f"[TrialRunStream] Failed to resolve frameworkRoot: {exc}")
                    yield _format_sse({"phase": "error", "error": f"Unable to resolve frameworkRoot: {exc}"})
                    return
                # Optional testmanager update prior to run
                if req.updateTestManager and (req.scenario or "").strip():
                    logger.info(f"[TrialRunStream] Updating testmanager for scenario: {req.scenario}")
                    try:
                        from ...services.config_service import update_test_manager_entry as _upd
                        upd = _upd(
                            root,
                            scenario=(req.scenario or "").strip(),
                            execute_value="Yes",
                            create_if_missing=True,
                            datasheet=(req.datasheet or None),
                            reference_id=(req.referenceId or None),
                            id_name=(req.idName or None),
                        )
                        if upd:
                            logger.info(f"[TrialRunStream] TestManager updated: {upd}")
                            yield _format_sse({"phase": "update", "update": upd})
                    except Exception as e:
                        logger.warning(f"[TrialRunStream] TestManager update failed: {e}")
                        pass
                test_dir = _detect_test_dir(root)
                test_dir.mkdir(parents=True, exist_ok=True)
                logger.info(f"[TrialRunStream] Test directory: {test_dir}")
                
                # Adapt spec content for trial run (replace credentials, add waits)
                from ...trial_spec_adapter import adapt_spec_content_for_trial
                adapted_content, was_adapted = adapt_spec_content_for_trial(req.testFileContent, root)
                if was_adapted:
                    logger.info(f"[TrialRunStream] Spec content adapted for trial run")
                    content = adapted_content
                else:
                    logger.info(f"[TrialRunStream] No adaptation needed, using original content")
                    content = req.testFileContent
                
                # Unskip for trial stream as well
                content, replaced = _unskip_tests_for_trial(content)
                logger.info(f"[TrialRunStream] Content length: {len(content)}, unskipped: {replaced}")
                with tempfile.NamedTemporaryFile(delete=False, suffix=".spec.ts", dir=str(test_dir)) as tmp:
                    tmp.write(content.encode("utf-8"))
                    tmp_path = tmp.name
                logger.info(f"[TrialRunStream] Created temp file: {tmp_path}")
                # Use path relative to framework root to avoid Windows path regex pitfalls
                try:
                    rel = _P(tmp_path).relative_to(root).as_posix()
                except ValueError:
                    rel = tmp_path.replace('\\', '/')
                logger.info(f"[TrialRunStream] Relative path: {rel}")
                
                # Determine ReferenceIDs for sequential multi-run (max 3)
                ref_ids: list[str] = []
                if req.referenceIds:
                    ref_ids = [r.strip() for r in req.referenceIds if r and str(r).strip()]
                elif req.referenceId and "," in req.referenceId:
                    ref_ids = [r.strip() for r in req.referenceId.split(",") if r.strip()]
                elif req.referenceId:
                    ref_ids = [req.referenceId.strip()]
                elif req.scenario:
                    excel_refs, excel_idname = _excel_refs(root, req.scenario)
                    if excel_refs:
                        ref_ids = excel_refs
                    if not req.idName and excel_idname:
                        req.idName = excel_idname
                if len(ref_ids) > 3:
                    ref_ids = ref_ids[:3]

                # Cap to maximum of 3 parallel runs
                if len(ref_ids) > 3:
                    ref_ids = ref_ids[:3]
                if not ref_ids:
                    ref_ids = [""]  # single run with no REFERENCE_ID

                # Launch separate browser process for each Reference ID
                import threading, queue as _queue
                base_env_overrides = trial_env_overrides(root)
                cmd_cwd = _resolve_playwright_command(rel, req.headed, project_root=root)
                base_cmd, base_cwd = cmd_cwd
                yield _format_sse({"phase": "prepared-parallel", "runs": len(ref_ids), "cmd": ' '.join(base_cmd), "cwd": base_cwd, "unskipped": replaced})

                events_q: _queue.Queue = _queue.Queue()
                success_map: dict[str, bool] = {}
                procs: list[tuple[str, subprocess.Popen]] = []

                # Start separate process for each Reference ID
                for idx, ref in enumerate(ref_ids):
                    run_label = (ref or f"run-{idx+1}")
                    trial_env = os.environ.copy()
                    if base_env_overrides:
                        trial_env.update(base_env_overrides)
                    if ref:
                        trial_env.update({
                            "REFERENCE_ID": ref,
                            "DATA_REFERENCE_ID": ref,
                        })
                    if req.idName:
                        trial_env.update({
                            "ID_NAME": req.idName,
                            "DATA_ID_NAME": req.idName,
                        })
                    cmd, cwd = _resolve_playwright_command(rel, req.headed, project_root=root)
                    logger.info(f"[TrialRunStream] Launching browser for {run_label}: {' '.join(cmd)}")
                    proc = subprocess.Popen(
                        cmd,
                        stdout=subprocess.PIPE,
                        stderr=subprocess.STDOUT,
                        text=True,
                        encoding="utf-8",
                        errors="replace",
                        cwd=str(root),
                        env=trial_env,
                    )
                    procs.append((run_label, proc))
                    events_q.put({"phase": "running", "referenceId": run_label})

                # Reader threads
                def _reader(label: str, p: subprocess.Popen):
                    try:
                        if p.stdout is None:
                            return
                        for line in iter(p.stdout.readline, ''):
                            if not line:
                                break
                            events_q.put({"phase": "chunk", "data": line.rstrip(), "referenceId": label})
                    except Exception as _e:
                        events_q.put({"phase": "chunk", "data": f"[reader-error] {label}: {_e}", "referenceId": label})
                    finally:
                        try:
                            rc = p.wait()
                            ok = (rc == 0)
                        except Exception:
                            ok = False
                        success_map[label] = ok
                        events_q.put({"phase": "done-single", "success": ok, "referenceId": label})

                threads: list[threading.Thread] = []
                for label, p in procs:
                    t = threading.Thread(target=_reader, args=(label, p), daemon=True)
                    t.start()
                    threads.append(t)

                # Pump queue to client until all done
                done_needed = len(procs)
                done_seen = 0
                while done_seen < done_needed or not events_q.empty():
                    try:
                        evt = await asyncio.get_event_loop().run_in_executor(None, events_q.get, )
                        if isinstance(evt, dict) and evt.get("phase"):
                            if evt.get("phase") == "done-single":
                                done_seen += 1
                            yield _format_sse(evt)
                    except Exception:
                        await asyncio.sleep(0.05)

                # Final summary
                overall_success = all(success_map.get(lbl, False) for (lbl, _) in procs) if procs else False
                yield _format_sse({"phase": "done", "success": overall_success, "runs": len(procs)})
            else:
                logger.info("[TrialRunStream] No frameworkRoot - using system temp")
                
                # Load trial credentials even without frameworkRoot
                from pathlib import Path as _PathLib
                project_root = _PathLib(__file__).resolve().parents[3]
                trial_env = os.environ.copy()
                env_overrides = trial_env_overrides(project_root)
                if env_overrides:
                    trial_env.update(env_overrides)
                    logger.info(f"[TrialRunStream] Added trial environment overrides: {list(env_overrides.keys())}")
                
                # Write temp spec file in system temp; rely on global PW config
                content, replaced = _unskip_tests_for_trial(req.testFileContent)
                with tempfile.NamedTemporaryFile(delete=False, suffix=".spec.ts") as tmp:
                    tmp.write(content.encode("utf-8"))
                    tmp_path = tmp.name
                logger.info(f"[TrialRunStream] Created temp file: {tmp_path}")
                cmd, cwd = _resolve_playwright_command(tmp_path, req.headed)
                logger.info(f"[TrialRunStream] Command: {' '.join(cmd)}, CWD: {cwd}, Headed: {req.headed}")
                yield _format_sse({"phase": "prepared", "headed": req.headed, "cmd": ' '.join(cmd), "cwd": cwd, "unskipped": replaced})
                logger.info(f"[TrialRunStream] Starting subprocess with headed={req.headed}")
                proc = subprocess.Popen(
                    cmd,
                    stdout=subprocess.PIPE,
                    stderr=subprocess.STDOUT,
                    text=True,
                    encoding="utf-8",
                    errors="replace",
                    cwd=cwd,
                    env=trial_env,  # Pass environment with trial credentials
                )
                logger.info(f"[TrialRunStream] Subprocess PID: {proc.pid}")

            # Single-run path for non-frameworkRoot
            yield _format_sse({"phase": "running"})
            logger.info("[TrialRunStream] Reading subprocess output...")
            assert proc.stdout is not None
            while True:
                line = await asyncio.get_event_loop().run_in_executor(None, proc.stdout.readline)
                if not line:
                    break
                yield _format_sse({"phase": "chunk", "data": line.rstrip()})
            ret = proc.wait()
            success = ret == 0
            logger.info(f"[TrialRunStream] Process finished with return code: {ret}, success: {success}")
            yield _format_sse({"phase": "done", "success": success})
        except Exception as exc:
            logger.error(f"[TrialRunStream] Error: {exc}", exc_info=True)
            yield _format_sse({"phase": "error", "error": str(exc)})
        finally:  # cleanup
            try:
                if 'tmp_path' in locals() and tmp_path and os.path.exists(tmp_path):
                    logger.info(f"[TrialRunStream] Cleaning up temp file: {tmp_path}")
                    os.unlink(tmp_path)
            except OSError as e:
                logger.warning(f"[TrialRunStream] Failed to cleanup temp file: {e}")
                pass

    return StreamingResponse(gen(), media_type="text/event-stream")


@router.get("/read-file")
async def read_file_from_repo(filePath: str, frameworkRoot: str | None = None):
    """Read the full content of a file from the framework repository.
    
    Args:
        filePath: Relative path to the file within the framework repository
        frameworkRoot: Optional framework root path/URL
        
    Returns:
        JSON with file content
    """
    try:
        # Resolve framework root
        root = resolve_framework_root(frameworkRoot) if frameworkRoot else resolve_framework_root()
        
        # Resolve target file path
        target = (root / filePath).resolve()
        
        # Security check: ensure path doesn't escape framework root
        root_resolved = root.resolve()
        if root_resolved not in target.parents and target != root_resolved:
            raise HTTPException(status_code=400, detail="File path escapes framework root")
        
        # Check if file exists
        if not target.exists():
            raise HTTPException(status_code=404, detail=f"File not found: {filePath}")
        
        if target.is_dir():
            raise HTTPException(status_code=400, detail="Path points to a directory, not a file")
        
        # Read file content
        try:
            content = target.read_text(encoding="utf-8", errors="replace")
        except Exception as read_error:
            raise HTTPException(status_code=500, detail=f"Failed to read file: {read_error}")
        
        return {
            "path": filePath,
            "content": content,
            "size": len(content),
            "lines": content.count('\n') + 1
        }
        
    except HTTPException:
        raise
    except Exception as exc:
        logger.error(f"Error reading file {filePath}: {exc}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Internal error: {exc}")

