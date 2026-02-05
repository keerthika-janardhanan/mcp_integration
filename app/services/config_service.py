from __future__ import annotations

import re
from pathlib import Path
from typing import Any, Dict, List, Optional


def _normalise_keyword(text: str) -> str:
    return re.sub(r"[^a-z0-9]+", " ", (text or "").strip().lower())


def _scenario_tokens(text: str) -> List[str]:
    return [token for token in re.findall(r"[a-z0-9]+", _normalise_keyword(text)) if len(token) >= 3]


def find_test_manager_path(framework_root: Path) -> Optional[Path]:
    direct = framework_root / "testmanager.xlsx"
    if direct.exists():
        return direct
    try:
        candidates = sorted(
            framework_root.glob("**/testmanager.xlsx"),
            key=lambda p: p.stat().st_mtime,
            reverse=True,
        )
    except Exception:
        candidates = list(framework_root.glob("**/testmanager.xlsx"))
    return candidates[0] if candidates else None


def update_test_manager_entry(
    framework_root: Path,
    scenario: str,
    execute_value: str = "Yes",
    create_if_missing: bool = True,
    datasheet: Optional[str] = None,
    reference_id: Optional[str] = None,
    id_name: Optional[str] = None,
    description_override: Optional[str] = None,
    allow_freeform_create: bool = False,
    ) -> Optional[Dict[str, Any]]:
    """Update or create a row in testmanager.xlsx using safe matching rules.

        Matching strategy (strict):
        - Match ONLY on TestCaseID column. We do not consider description for matching anymore.
        - Exact (case-insensitive) match first; if not found, allow cautious fuzzy match on ID tokens.
        - If still not found and creation is allowed, create a new row with TestCaseID set to the provided scenario
            and optionally use description_override for TestCaseDescription.
    """

    try:
        from openpyxl import load_workbook  # type: ignore
        from openpyxl import Workbook  # type: ignore
    except Exception:
        raise RuntimeError("openpyxl is required to update testmanager.xlsx")

    tm_path = find_test_manager_path(framework_root)
    if not tm_path:
        if not create_if_missing:
            return None
        # Create a fresh testmanager.xlsx with standard headers
        tm_path = framework_root / "testmanager.xlsx"
        try:
            wb_new = Workbook()
            ws_new = wb_new.active
            ws_new.title = "ExecutionPlan"
            # Standard column headers used across the framework
            headers = [
                "TestCaseID",
                "TestCaseDescription",
                "Execute",
                "DatasheetName",
                "ReferenceID",
                "IDName",
            ]
            for idx, h in enumerate(headers, start=1):
                ws_new.cell(row=1, column=idx, value=h)
            wb_new.save(tm_path)
        except Exception:
            return None

    try:
        wb = load_workbook(tm_path)
    except Exception:
        return None

    # Prefer a sheet named 'ExecutionPlan'; otherwise, pick the first sheet that appears
    # to contain the required headers (TestCaseDescription + Execute). Fallback to active sheet.
    ws = None
    try:
        if "ExecutionPlan" in wb.sheetnames:
            ws = wb["ExecutionPlan"]
    except Exception:
        ws = None
    if ws is None:
        best_sheet = None
        best_score = -1
        for sheet in wb.worksheets:
            try:
                header_cells = list(sheet[1]) if sheet.max_row >= 1 else []
            except Exception:
                continue
            header_map_probe: Dict[str, int] = {}
            for idx, cell in enumerate(header_cells, start=1):
                value = cell.value
                if value is None:
                    continue
                header_map_probe[_normalise_keyword(str(value))] = idx

            def _has(col_name: str) -> bool:
                key = _normalise_keyword(col_name)
                return any(key == hk or key in hk for hk in header_map_probe.keys())

            score = 0
            if _has("TestCaseDescription") or _has("Scenario") or _has("Description"):
                score += 1
            if _has("Execute") or _has("Run") or _has("Enabled"):
                score += 1
            if _has("DatasheetName"):
                score += 1
            if score > best_score:
                best_score = score
                best_sheet = sheet
        if best_sheet is not None and best_score >= 2:  # require at least desc+execute
            ws = best_sheet
    if ws is None:
        ws = wb.active
    header_map: Dict[str, int] = {}
    for idx, cell in enumerate(ws[1], start=1):
        value = cell.value
        if value is None:
            continue
        header_map[_normalise_keyword(str(value))] = idx

    def _find_column(*candidates: str) -> Optional[int]:
        for candidate in candidates:
            key = _normalise_keyword(candidate)
            for header_key, col_idx in header_map.items():
                if key == header_key or key in header_key:
                    return col_idx
        return None

    desc_col = _find_column("TestCaseDescription", "Scenario", "Description")
    execute_col = _find_column("Execute", "Run", "Enabled")
    id_col = _find_column("TestCaseID", "ID", "Identifier")
    datasheet_col = _find_column("DatasheetName", "DataSheet", "Data Sheet")
    reference_col = _find_column("ReferenceID", "Reference Id", "Reference")
    idname_col = _find_column("IDName", "IdentifierName", "RowIdentifier")

    if not desc_col or not execute_col:
        return None

    scenario_norm = _normalise_keyword(scenario)
    scenario_tokens = set(_scenario_tokens(scenario))

    # Always treat the provided scenario as intended TestCaseID for matching/creation
    scenario_stripped = (scenario or "").strip()
    is_id_like = True

    matched_row = None
    matched_description = None
    best_row = None
    best_score = 0

    # First pass: exact, case-insensitive
    for row_idx in range(2, ws.max_row + 1):
        case_id = ""
        if id_col:
            case_id = str(ws.cell(row=row_idx, column=id_col).value or "").strip()
        if not case_id or not scenario_stripped:
            continue
        if case_id.lower() == scenario_stripped.lower():
            matched_row = row_idx
            matched_description = case_id
            break

    # Second pass: fuzzy matching with stricter thresholds
    # If this is a non-ID-like natural language scenario AND caller explicitly allows freeform creation,
    # we deliberately SKIP fuzzy matching so that we don't hijack an existing row that shares generic tokens
    # like 'create', 'team', 'project'. This ensures a new row is created unless there is an exact
    # description match (handled above).
    if matched_row is None:
        for row_idx in range(2, ws.max_row + 1):
            case_id = ""
            if id_col:
                case_id = str(ws.cell(row=row_idx, column=id_col).value or "").strip()
            if not case_id:
                continue
            id_norm = _normalise_keyword(case_id)
            id_tokens = set(_scenario_tokens(case_id))
            # Fuzzy only on ID
            if scenario_norm and id_norm and (scenario_norm in id_norm or id_norm in scenario_norm):
                matched_row = row_idx
                matched_description = case_id
                break
            if scenario_tokens:
                score = len(scenario_tokens & id_tokens)
                if score > best_score:
                    best_score = score
                    best_row = row_idx
                    matched_description = case_id

    # Accept only if overlap is meaningful (>= 2 tokens)
    if matched_row is None and best_row is not None and best_score >= 2:
        matched_row = best_row

    rel_path = str(tm_path.relative_to(framework_root)).replace("\\", "/")

    if matched_row is not None:
        exec_cell = ws.cell(row=matched_row, column=execute_col)
        previous_value = str(exec_cell.value).strip() if exec_cell.value is not None else ""
        changed = False
        if execute_value is not None and previous_value != execute_value:
            exec_cell.value = execute_value
            changed = True
        # Optionally update description if provided
        if description_override and desc_col:
            cell = ws.cell(row=matched_row, column=desc_col)
            if (str(cell.value or "").strip() != str(description_override).strip()):
                cell.value = description_override
                changed = True
        if datasheet_col and datasheet is not None:
            cell = ws.cell(row=matched_row, column=datasheet_col)
            if (str(cell.value or "").strip() != str(datasheet).strip()):
                cell.value = datasheet
                changed = True
        if reference_col and reference_id is not None:
            cell = ws.cell(row=matched_row, column=reference_col)
            if (str(cell.value or "").strip() != str(reference_id).strip()):
                cell.value = reference_id
                changed = True
        if idname_col and id_name is not None:
            cell = ws.cell(row=matched_row, column=idname_col)
            if (str(cell.value or "").strip() != str(id_name).strip()):
                cell.value = id_name
                changed = True
        if changed:
            try:
                wb.save(tm_path)
            except Exception:
                pass
            return {
                "path": rel_path,
                "mode": "updated",
                "description": matched_description or scenario,
                "previous": previous_value,
                "execute": execute_value,
                "matched_row": matched_row,
                "matched_description": matched_description,
                "datasheet_applied": datasheet if datasheet_col else None,
                "reference_applied": reference_id if reference_col else None,
                "idname_applied": id_name if idname_col else None,
            }
        return {
            "path": rel_path,
            "mode": "unchanged",
            "description": matched_description or scenario,
            "previous": previous_value,
            "execute": execute_value,
            "matched_row": matched_row,
            "matched_description": matched_description,
            "datasheet_applied": datasheet if datasheet_col else None,
            "reference_applied": reference_id if reference_col else None,
            "idname_applied": id_name if idname_col else None,
        }

    if not create_if_missing:
        return None

    # Only auto-create for ID-like input by default; can be overridden.
    # Creation allowed, we will create a row with the provided scenario as TestCaseID

    new_row = ws.max_row + 1
    scenario_text = scenario.strip()
    # If provided, use description_override for description column
    ws.cell(row=new_row, column=desc_col).value = (description_override or scenario_text)
    ws.cell(row=new_row, column=execute_col).value = execute_value
    if id_col:
        ws.cell(row=new_row, column=id_col).value = scenario_text
    if datasheet_col and datasheet is not None:
        ws.cell(row=new_row, column=datasheet_col).value = datasheet
    if reference_col and reference_id is not None:
        ws.cell(row=new_row, column=reference_col).value = reference_id
    if idname_col and id_name is not None:
        ws.cell(row=new_row, column=idname_col).value = id_name
    try:
        wb.save(tm_path)
    except Exception:
        pass
    return {
        "path": rel_path,
        "mode": "created",
        "description": scenario_text,
        "previous": "",
        "execute": execute_value,
        "matched_row": ws.max_row,
        "matched_description": scenario_text,
        "datasheet_applied": datasheet if datasheet_col else None,
        "reference_applied": reference_id if reference_col else None,
        "idname_applied": id_name if idname_col else None,
    }
