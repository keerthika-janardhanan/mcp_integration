import io
import zipfile
import xml.etree.ElementTree as ET
from typing import Optional

import pandas as pd
from openpyxl import load_workbook


XL_NS = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}"
_SANITIZE_TAGS = (
    f"{XL_NS}autoFilter",
    f"{XL_NS}dataValidations",
    f"{XL_NS}conditionalFormatting",
)


def _strip_problematic_nodes(xml_bytes: bytes) -> bytes:
    """Remove Excel worksheet nodes that embed filter/validation formulas."""
    try:
        root = ET.fromstring(xml_bytes)
    except ET.ParseError:
        return xml_bytes

    removed = False
    for tag in _SANITIZE_TAGS:
        for node in list(root.findall(tag)):
            root.remove(node)
            removed = True

    if not removed:
        return xml_bytes

    cleaned = ET.tostring(root, encoding="utf-8")
    return b'<?xml version="1.0" encoding="UTF-8" standalone="yes"?>' + cleaned


def _sanitize_excel_bytes(raw: bytes) -> bytes:
    """Remove worksheet nodes that can trip regex parsing while preserving workbook structure."""
    try:
        source = io.BytesIO(raw)
        cleaned_buffer = io.BytesIO()
        with zipfile.ZipFile(source) as zin, zipfile.ZipFile(cleaned_buffer, "w") as zout:
            for item in zin.infolist():
                data = zin.read(item.filename)
                if item.filename.startswith("xl/worksheets/") and item.filename.endswith(".xml"):
                    data = _strip_problematic_nodes(data)
                zout.writestr(item, data)
        return cleaned_buffer.getvalue()
    except zipfile.BadZipFile:
        # Not an XLSX file (likely XLS); return original bytes.
        return raw
    except Exception:
        return raw


def _read_excel_from_bytes(data: bytes, **kwargs) -> pd.DataFrame:
    buffer = io.BytesIO(data)
    return pd.read_excel(buffer, **kwargs)


def _read_excel_with_openpyxl(data: bytes) -> pd.DataFrame:
    """Fallback reader that ignores Excel metadata entirely."""
    buffer = io.BytesIO(data)
    wb = load_workbook(buffer, data_only=True, read_only=False)
    try:
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
    finally:
        wb.close()

    if not rows:
        return pd.DataFrame()

    header = ["" if cell is None else str(cell) for cell in rows[0]]
    width = len(header)
    data_rows = []
    for row in rows[1:]:
        normalized = []
        for idx in range(width):
            value = row[idx] if idx < len(row) else None
            normalized.append("" if value is None else value)
        data_rows.append(normalized)

    return pd.DataFrame(data_rows, columns=header)


def load_excel_template(uploaded_file, dtype: Optional[type] = None) -> pd.DataFrame:
    """
    Load an uploaded Excel template into a DataFrame while stripping Excel artefacts
    that can introduce invalid regular expressions (e.g., custom filters).
    """
    if uploaded_file is None:
        raise ValueError("No template file provided.")

    # Streamlit's UploadedFile exposes getvalue(); fall back to read/seek for other-like objects.
    if hasattr(uploaded_file, "getvalue"):
        raw_bytes = uploaded_file.getvalue()
    else:
        raw_bytes = uploaded_file.read()
        if hasattr(uploaded_file, "seek"):
            uploaded_file.seek(0)

    cleaned_bytes = _sanitize_excel_bytes(raw_bytes)
    read_kwargs = {"dtype": dtype} if dtype else {}

    try:
        df = _read_excel_from_bytes(cleaned_bytes, **read_kwargs)
    except Exception:
        try:
            df = _read_excel_from_bytes(raw_bytes, **read_kwargs)
        except Exception:
            df = _read_excel_with_openpyxl(cleaned_bytes)

    return df
